from xml.dom import minidom
from xlutils.copy import copy
from openpyxl import Workbook, load_workbook

import xlwt
import xlrd
import os
import gc

# -- Inicialización del programa -- #

## Variables Generales del programa ##
matchID=""
converts=[]
derivadas=[]
merges=[]
lookups=[]
convertCount=0
derivCount=0
lookupsCount=0
inputConverts=[]
outputConverts=[]
inputMerges=[]
outputMerges=[]
inputLookups=[]
outputLookups=[]
outputDerivs=[]
inputDerivs=[]
nombresHojas=[]
nombreConexion=""
nombreHoja=""

    
## Cursores de Escritura en Excel ##
cursorVertical=0
cursorAuxiliar=0
cursorVertical2=0
cursorAuxiliar2=0
cursorHorizontal=4


actividad=""
proceso=""
tipoCaja=""
nombreCaja=""
campoFuente=""
campoDestino=""
parejas=[]


    
## Campos Predefinidos de escritura ##
def resetCampos():
    global tipoCaja
    tipoCaja=""
    global nombreCaja
    nombreCaja=""
    global campoFuente
    campoFuente=""
    global campoDestino
    campoDestino=""
    return



## Leer directorio de archivos ##
listadoArchivos = os.listdir('.')
listadoLimpio=[]

listadoArchivos.reverse()


## Excluir Archivos no deseados ##
for each in listadoArchivos:
    if ".dtsx" in each: #and "DimEstructura" in each:     
        listadoLimpio.append(each)

#listadoLimpio=["ExtraccionDimProducto.dtsx"]


## Crear o Abrir Excel ##


## Ingresar Registro en Excel ##
def escribirRegistro():
    global cursorVertical
    global actividad
    global proceso
    global tipoCaja
    global nombreCaja
    global campoFuente
    global campoDestino
    #global hoja

    hoja.write(cursorVertical,0,actividad)
    hoja.write(cursorVertical,1,proceso)
    hoja.write(cursorVertical,2,tipoCaja)
    hoja.write(cursorVertical,3,nombreCaja)
    hoja.write(cursorVertical,4,campoFuente)
    hoja.write(cursorVertical,5,campoDestino)
    cursorVertical+=1
    return


## Crear Encabezados ##
actividad="Actividad"
proceso="Proceso"
tipoCaja="Tipo Caja"
nombreCaja="Nombre Caja"
campoFuente="Campo Fuente"
campoDestino="Campo Destino"
tablaDestino=""




# -- Ejecución del programa -- #

## Procesar Archivo ##
for archivo in listadoLimpio:
    
    cursorVertical=0
    cursorAuxiliar=0
    cursorVertical2=0
    cursorAuxiliar2=0
    cursorHorizontal=4
    matchID=""
    converts=[]
    derivadas=[]
    merges=[]
    lookups=[]
    convertCount=0
    derivCount=0
    lookupsCount=0
    inputConverts=[]
    outputConverts=[]
    inputMerges=[]
    outputMerges=[]
    inputLookups=[]
    outputLookups=[]
    inputPivots=[]
    outputPivots=[]
    pivots=[]
    outPuts=[]
    conections=[]
    pivotCount=0
    outputDerivs=[]
    nombrePivot=""
    inputDerivs=[]
    nombresHojas=[]
    sorts=[]
    sourceBD=""
    inputSorts=[]
    outputSorts=[]
    conectionManagers=[]
    subConectionManagers=[]
    nombreConexion=""
    path=""
    
    xmldoc = minidom.parse(archivo)
    tempArch=archivo
    
    
    ## Escribir nombre de archivo valido como nombre de hoja en Excel ##
    if ("Extraccion" in tempArch):
        libro = xlwt.Workbook()
        archivo=archivo[10:-5]
        hoja = libro.add_sheet(archivo)
        escribirRegistro()
        resetCampos()
        actividad="Extraccion"
    else:
        try:
            archivo=archivo[5:-5]
            libroAux = xlrd.open_workbook(os.path.join("C:/Users/Usuario/Desktop/Nalsani SSIS/Nalsani SSIS/"+archivo+"/", archivo+".xls"))
            libro = copy(libroAux)
            hoja =  libro.get_sheet(0)
            cursorVertical=libroAux.sheet_by_index(0).nrows
            #escribirRegistro()
            resetCampos()
            actividad="Carga"
            print("Aparentemente entre")
        except Exception as e:
            print(str(e))
            libro = xlwt.Workbook()
            hoja=libro.add_sheet(archivo)
            escribirRegistro()
            resetCampos()
            actividad="Carga"

    

        
    print("Procesando Archivo: "+tempArch)
    ## Obtener Conexiones ##
    if actividad=="Extraccion":
        path1=""
        conectionManagers = xmldoc.getElementsByTagName('DTS:ConnectionManager')
        for conectionManager in conectionManagers:
            if conectionManager.hasAttribute("DTS:refId"):
                subConectionManagers = conectionManager.getElementsByTagName('DTS:ConnectionManager')
                for subConectionManager in subConectionManagers:
                    path = subConectionManager.attributes['DTS:ConnectionString'].value.split(";")
                    for miniPath in path:
                        #print("MP: "+miniPath)
                        if ("Data Source" in miniPath):
                            path1=miniPath[12:]
                            break
                refId = conectionManager.attributes['DTS:refId'].value
                nombre = conectionManager.attributes['DTS:ObjectName'].value
                conections.append([refId,path1])
                #actividad="Extraccion"
                #proceso="Admin Conexiones"
                #tipoCaja="Admin Conexiones"
                #nombreCaja="Conexion"
                #campoFuente=nombre
                #campoDestino=path
                #escribirRegistro()

            
    ## Obtener Raiz ##
    SuperDTS = xmldoc.getElementsByTagName('DTS:Executable')
    for DTS in SuperDTS:
        
        ## Extracción Data Flow Task ##
        if (DTS.hasAttribute("DTS:Description") and "Data Flow Task" == DTS.attributes['DTS:Description'].value and "Cache" not in DTS.attributes['DTS:ObjectName'].value   ):

            ## Definir Campo "Proceso" ##
            proceso = DTS.attributes['DTS:ObjectName'].value
            
            #print("Proceso: "+proceso)
            OLE_DBS = DTS.getElementsByTagName('component')

            
            ## Extracción OLEDBDestination ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.OLEDBDestination" == OLE_DB.attributes['componentClassID'].value and "Rechazos" not in OLE_DB.attributes['name'].value):

                    ## Definir Campo "Tipo Caja" ##
                    tipoCaja="Destino"
        
                    ## Definir Campo "Nombre Caja" ##
                    nombreCaja = OLE_DB.attributes['name'].value

                    inputs = OLE_DB.getElementsByTagName('inputColumn')
                    if (actividad=="Carga"):
                        properties = OLE_DB.getElementsByTagName('property')
                        for properti in properties:
                            if (properti.attributes['name'].value == "OpenRowset"):
                                tablaDestino = properti.firstChild.nodeValue
                                #print("tablaDestino: "+tablaDestino)
                                
                    for inputColumn in inputs:
                        
                        ## Definir Campo "Campo Fuente" ##
                        campoFuente=inputColumn.attributes['cachedName'].value
                        tipoFuente=inputColumn.attributes['cachedDataType'].value
                        matchID=inputColumn.attributes['externalMetadataColumnId'].value
                        outputs = OLE_DB.getElementsByTagName('externalMetadataColumn')

                        ## Buscar Columna Correspondiente ##
                        for outputColumn in outputs:
                            
                            if (outputColumn.attributes['refId'].value==matchID):
                                
                                ## Definir Campo "Campo Destino" ##
                                tipoCaja="Destino"
                                campoDestino = outputColumn.attributes['name'].value
                                tipoDestino = outputColumn.attributes['dataType'].value
                                if tipoFuente != tipoDestino:
                                    parejas.append([tempArch,nombreCaja,campoFuente,tipoFuente,campoDestino,tipoDestino])
                                escribirRegistro()
                                tipoCaja="Tabla Destino"
                                campoDestino=tablaDestino
                                escribirRegistro()
                                
            resetCampos()
            ## Extracción DataConvert ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.DataConvert" == OLE_DB.attributes['componentClassID'].value):

                    nombreCaja=OLE_DB.attributes['name'].value
                    converts.append(nombreCaja)
                    #print(converts)
                    inputConverts += OLE_DB.getElementsByTagName('inputColumn')
                    outputConverts += OLE_DB.getElementsByTagName('outputColumn')
            ## Extracción Sort ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.Sort" == OLE_DB.attributes['componentClassID'].value):

                    nombreCaja=OLE_DB.attributes['name'].value
                    sorts.append(nombreCaja)
                    #print(converts)
                    inputSorts += OLE_DB.getElementsByTagName('inputColumn')
                    outputSorts += OLE_DB.getElementsByTagName('outputColumn')
                    

            resetCampos()
            ## Extracción MergeJoin ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.MergeJoin" == OLE_DB.attributes['componentClassID'].value):

                    nombreCaja=OLE_DB.attributes['name'].value
                    merges.append(nombreCaja)
                    #print(converts)
                    inputMerges += OLE_DB.getElementsByTagName('inputColumn')
                    outputMerges += OLE_DB.getElementsByTagName('outputColumn')

            resetCampos()
            ## Extracción DerivedColumn ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.DerivedColumn" == OLE_DB.attributes['componentClassID'].value):
                    nombreDerivada=OLE_DB.attributes['name'].value
                    derivadas.append(nombreDerivada)
                    inputDerivs += OLE_DB.getElementsByTagName('inputColumn')
                    outputDerivs += OLE_DB.getElementsByTagName('outputColumn')

            ## Extracción Excel ##        
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.ExcelSource" == OLE_DB.attributes['componentClassID'].value):
                    outPuts = OLE_DB.getElementsByTagName('output')
                    properties = OLE_DB.getElementsByTagName('property')
                    connectionRefID = OLE_DB.getElementsByTagName('connection')[0].attributes['connectionManagerRefId'].value
                    #print("CRIF: "+connectionRefID)
                    for conexion in conections:
                        #conexion[0]=RefID
                        #conexion[1]=path
                        #print("con0: "+conexion[0])
                        
                        if (conexion[0]==connectionRefID):
                            #print("MAtch")
                            nombreConexion=conexion[1]
                        
                    for properti in properties:
                        if (properti.hasAttribute("name") and "OpenRowset" == properti.attributes['name'].value):
                            nombreHoja = properti.firstChild.nodeValue[:-1]
                    for outPut in outPuts:
                        if "Error" not in outPut.attributes['name'].value:
                            outputColumns = outPut.getElementsByTagName('outputColumn')
                            for outputColumn in outputColumns:
                                
                                externalID = outputColumn.attributes['externalMetadataColumnId'].value
                                externalColumns = outPut.getElementsByTagName('externalMetadataColumn')
                                for externalColumn in externalColumns:
                                    refId = externalColumn.attributes['refId'].value
                                    #print("ext - "+externalID)
                                    #print("ref - "+refId)
                                    if externalID==refId:
                                        campoFuente=externalColumn.attributes['name'].value
                                        campoDestino=outputColumn.attributes['name'].value
                                        nombreCaja=OLE_DB.attributes['name'].value
                                        tipoCaja="Excel"
                                        escribirRegistro()
                                        
                                        campoFuente=nombreHoja
                                        tipoCaja="Excel Source Table"
                                        escribirRegistro()
                                        
                                        campoFuente=nombreConexion
                                        tipoCaja="Excel Source Connection"
                                        escribirRegistro()

            ## Extracción Lookups ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.Lookup" == OLE_DB.attributes['componentClassID'].value):
                    nombreLookup=OLE_DB.attributes['name'].value
                    lookups.append(nombreLookup)
                    inputLookups = OLE_DB.getElementsByTagName('inputColumn')
                    outputLookups = OLE_DB.getElementsByTagName('outputColumn')
                    for outputLookup in outputLookups:
                        propertiesOut = outputLookup.getElementsByTagName('property')
                        for propertiOut in propertiesOut:
                            if ("System.String" in propertiOut.attributes['dataType'].value ):
                                campoDestino=propertiOut.firstChild.nodeValue
                                for inputLookup in inputLookups:
                                    propertiesIn = inputLookup.getElementsByTagName('property')
                                    for propertiIn in propertiesIn:
                                        ## Definir Campo Destino##
                                        if ("System.String" in propertiIn.attributes['dataType'].value ):
                                            #campoFuente=propertiIn.firstChild.nodeValue
                                            campoFuente=inputLookup.attributes['lineageId'].value.split("[")
                                            campoFuente=campoFuente[len(campoFuente)-1][:-1]
                                            nombreCaja = OLE_DB.attributes['name'].value
                                            tipoCaja="LookUp Column"
                                                                
                                            lookupsCount+=1
                                            escribirRegistro()
                                            resetCampos()
            ## Extracción Pivots ##
            for OLE_DB in OLE_DBS:
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.Pivot" == OLE_DB.attributes['componentClassID'].value):
                    nombrePivot=OLE_DB.attributes['name'].value
                    #lookups.append(nombreLookup)
                    outPuts = OLE_DB.getElementsByTagName('output')
                    for outPut in outPuts:
                        if outPut.attributes['name'].value=="Pivot Default Output":
                            outputPivots = OLE_DB.getElementsByTagName('outputColumn')
                            for outputPivot in outputPivots:
                                propertiesOut = outputPivot.getElementsByTagName('property')
                                for propertiOut in propertiesOut:
                                    if ("PivotKeyValue" in propertiOut.attributes['name'].value ):
                                        try:
                                            campoFuente=propertiOut.firstChild.nodeValue
                                            campoDestino=outputPivot.attributes['name'].value
                                            tipoCaja="Pivot Column"
                                            nombreCaja=nombrePivot
                                            escribirRegistro()
                                            resetCampos()
                                        except:
                                            pass

            resetCampos()
            
            ## OLEDBSource ##
            for OLE_DB in OLE_DBS:        
                if (OLE_DB.hasAttribute("componentClassID") and "Microsoft.OLEDBSource" == OLE_DB.attributes['componentClassID'].value and "HOLDING" not in OLE_DB.attributes['name'].value):
                    if (actividad=="Extraccion"):
                        try:
                            connection = OLE_DB.getElementsByTagName('connection')[0]
                            sourceBD = connection.attributes['connectionManagerRefId'].value.split("[")[1].split("]")[0]
                        except:
                            sourceBD="Invalid"
                            #print(connection.attributes['connectionManagerRefId'].value)
                    outputs = OLE_DB.getElementsByTagName('output')
                    for output in outputs:
                        if (output.hasAttribute("name") and "OLE DB Source Error Output" != output.attributes['name'].value):
                            inputs = output.getElementsByTagName('outputColumn')
                            for inputColumn in inputs:
                                
                                campoDestino=inputColumn.attributes['name'].value
                                tipoDestino=inputColumn.attributes['dataType'].value
                                matchID=inputColumn.attributes['externalMetadataColumnId'].value
                                lineagesID=inputColumn.attributes['lineageId'].value
                                outputs = OLE_DB.getElementsByTagName('externalMetadataColumn')
                                
                                
                                    
                                for outputColumn in outputs:
                                    if (outputColumn.attributes['refId'].value==matchID):
                                        campoFuente = outputColumn.attributes['name'].value
                                        tipoFuente = outputColumn.attributes['dataType'].value
                                        tipoCaja="Origen"
                                        nombreCaja = OLE_DB.attributes['name'].value
                                        if (tipoFuente!=tipoDestino):
                                            parejas.append([tempArch,nombreCaja,campoFuente,tipoFuente,campoDestino,tipoDestino])
                                       # print("Co1lumna Destino: "+columnaDestino)
                                        escribirRegistro()
                                        if actividad=="Extraccion":
                                            tipoCaja="Conexion Origen"
                                            #print(sourceBD)
                                            #print(campoFuente+" -> "+sourceBD)
                                            campoFuente=sourceBD
                                            #print(sourceBD)
                                            escribirRegistro()
                                            resetCampos()
                                
                                

                                    
            ## Data Conversion ##
            #print("Finalizando Origenes")        
            for inputConvert in inputConverts:
                lineagesIDX = inputConvert.attributes['lineageId'].value                                       
                for outputConvert in outputConverts:
                       if ( "Audit" not in inputConvert.attributes['lineageId'].value and "ErrorCode" not in outputConvert.attributes['name'].value):
                           properties = outputConvert.getElementsByTagName('property')
                           for properti in properties:
                               destinationLineage = properti.firstChild.nodeValue.split("{")
                               for desti in destinationLineage:
                                   if ("Package" in desti):
                                       destinationLineage=desti.split("}")[0]
                                       break
                               if (destinationLineage == lineagesIDX):
                            
                                   tipoCaja="Data Conversion"
                                   ## Definir Campos de Impresion Data Conversion##
                                   
                                   nombreCaja=inputConvert.parentNode.parentNode.parentNode.parentNode.attributes['name'].value

                                   campoFuente=inputConvert.attributes['cachedName'].value
                                   campoDestino=outputConvert.attributes['name'].value
                                   convertCount+=1

                                   escribirRegistro()
                                   resetCampos()
            ## Sort ##
            #print("Finalizando Origenes")        
            for inputSort in inputSorts:
                lineagesIDX = inputSort.attributes['lineageId'].value                                       
                for outputSort in outputSorts:
                       if ( "Audit" not in inputSort.attributes['lineageId'].value and "ErrorCode" not in outputSort.attributes['name'].value):
                           properties = outputSort.getElementsByTagName('property')
                           for properti in properties:
                               destinationLineage = properti.firstChild.nodeValue.split("{")
                               for desti in destinationLineage:
                                   if ("Package" in desti):
                                       destinationLineage=desti.split("}")[0]
                                       break
                               if (destinationLineage == lineagesIDX):
                            
                                   tipoCaja="Sort"
                                   ## Definir Campos de Impresion Data Conversion##
                                   
                                   nombreCaja=inputSort.parentNode.parentNode.parentNode.parentNode.attributes['name'].value

                                   campoFuente=inputSort.attributes['cachedName'].value
                                   campoDestino=outputSort.attributes['name'].value
                                   convertCount+=1

                                   escribirRegistro()
                                   resetCampos()
            ## Merge Join ##
            #print("Finalizando Data Conversion")        
            for inputMerge in inputMerges:
                lineagesIDX = inputMerge.attributes['refId'].value                                       
                for outputMerge in outputMerges:
                       if ( "Audit" not in inputMerge.attributes['lineageId'].value and "ErrorCode" not in outputMerge.attributes['name'].value):
                           properties = outputMerge.getElementsByTagName('property')
                           for properti in properties:
                               destinationLineage = properti.firstChild.nodeValue.split("{")
                               for desti in destinationLineage:
                                   if ("Package" in desti):
                                       destinationLineage=desti.split("}")[0]
                                       break
                               
                               if (destinationLineage == lineagesIDX):
                                   #print(destinationLineage)
                            
                                   tipoCaja="Merge Join"
                                   ## Definir Campos de Impresion Data Conversion##
                                   
                                   nombreCaja=inputMerge.parentNode.parentNode.parentNode.parentNode.attributes['name'].value

                                   campoFuente=inputMerge.attributes['cachedName'].value
                                   campoDestino=outputMerge.attributes['name'].value
                                   #print("Merge Succesful")
                                   #print(campoFuente)
                                   #print(campoDestino)

##                                                   carac1Nombre="DataType Entrada"
##                                                   carac2Nombre="DataType Salida"
##                                                   carac3Nombre="Length Entrada"
##                                                   carac4Nombre="Length Salida"
##
##                                                   carac1Valor=inputConvert.attributes['cachedDataType'].value
##                                                   carac2Valor=outputConvert.attributes['dataType'].value
##                                                   carac3Valor=inputConvert.attributes['cachedLength'].value
##                                                   carac4Valor=outputConvert.attributes['length'].value

                                   escribirRegistro()
                                   resetCampos()

                ## Derived Column ##
            #print("Finalizando Columnas Merge joins")
            for inputDeriv in inputDerivs:
                lineagesIDX = inputDeriv.attributes['lineageId'].value                                       
                for outputDeriv in outputDerivs:
                       properties = outputDeriv.getElementsByTagName('property')
                       for properti in properties:
                           destinationLineage = properti.firstChild.nodeValue.split("{")
                           for desti in destinationLineage:
                               if ("Package" in desti):
                                   destinationLineage=desti.split("}")[0]
                                   break
                          
                           if (destinationLineage == lineagesIDX):
                               ## Definir Campos de Impresion Data Conversion##

                               nombreCaja=inputDeriv.parentNode.parentNode.parentNode.parentNode.attributes['name'].value
                               tipoCaja="Derived Column"
                               campoFuente=inputDeriv.attributes['cachedName'].value
                               campoDestino=outputDeriv.attributes['name'].value
                              #print(outputDeriv.attributes['refId'].value)
                               carac1Nombre="Expresión"
                               for properti in properties:
                                   if ("Friendly" in properti.attributes['description'].value):
                                       carac1Valor = properti.firstChild.nodeValue
                               derivCount+=1
                               escribirRegistro()
                               resetCampos()
            #print("Finalizando Columnas Derivadas")
    
            
    try:
        libro.save(os.path.join("C:/Users/Usuario/Desktop/Nalsani SSIS/Nalsani SSIS/"+archivo+"/", archivo+".xls"))
    except:
        libro.save(archivo+".xls")
    gc.collect()
print("Finalizando procesamiendo\n")

##book = xlwt.Workbook()
##sheet = book.add_sheet("ListaConversiones")
##contx=0
##conty=0
##([tempArch,nombreCaja,campoFuente,tipoFuente,campoDestino,tipoDestino])
##sheet.write(0,0,"Archivo")
##sheet.write(0,1,"Nombre Caja")
##sheet.write(0,2,"Campo Fuente")
##sheet.write(0,3,"Tipo Fuente")
##sheet.write(0,4,"Campo Destino")
##sheet.write(0,5,"Tipo Destino")
##for pareja in parejas:
##    conty+=1
##    contx=0
##    for campo in pareja:
##        sheet.write(conty,contx,campo)
##        contx+=1
##book.save("Listado Conversion Bugs.xls")
        

        
#           parejas.append([nombreCaja,campoFuente,tipoFuente,campoDestino,tipoDestino])

