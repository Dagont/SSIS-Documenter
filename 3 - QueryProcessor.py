from xml.dom import minidom
from xlutils.copy import copy 
import xlwt
import xlrd
import gc
import os
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
    # first open using xlrd
    #print("Filename: "+filename)
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active
    #sheet1 = book1.get_active_sheet()

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1


## Leer directorio de archivos ##
listadoArchivos = os.listdir(os.getcwd())
listadoLimpio=[]
diccionarioTablas=[]
nombreOrigen=""
nombreAlias=""
tablaOrigen=""
nombreTabla=""
nombreAlias1=""

#print(listadoArchivos)


## Excluir Archivos no deseados ##
for each in listadoArchivos:
    if "Extraccion" in each:# and "DimCliente" in each:
        listadoLimpio.append(each)

        
for archivo in listadoLimpio:
    cont=0
    tablaPrincipal=""
    diccionarioTablas=[]
    nombreOrigen=""
    nombreAlias=""
    tablaOrigen=""
    nombreTabla=""
    nombreAlias1=""
    for root, dirs, files in os.walk(archivo[10:-5], topdown=False):
        print("Procesando: "+root)
        contieneXML=False
        for name in files:
            if (".xml" in name):
                contieneXML=True
                #print("ROOT: "+root)
                #print("Procesando: "+name)
                xmldoc = minidom.parse(os.path.join(root, name))
                ##FROM##
                froms = xmldoc.getElementsByTagName('from_clause')
                #print("## Imprimiendo FROM ##")
                for fromx in froms:
                    tableReferences = fromx.getElementsByTagName('named_table_reference')
                    for tableReference in tableReferences:
                          nombreTabla=""
                          nombreAlias1=""
                          #print("--")
                          tableNames=tableReference.getElementsByTagName('table_name')
                          for tableName in tableNames:
                              full_name=tableReference.getElementsByTagName('full_name')[0]
                              nombreTabla = full_name.firstChild.nodeValue
                              if not (tablaPrincipal):
                                  tablaPrincipal=nombreTabla
                              #print("Nombre Origen: "+nombreTabla)
                          try:
                              aliasNames=tableReference.getElementsByTagName('alias_clause')
                              for aliasName in aliasNames:
                                  object_names=aliasName.getElementsByTagName('object_name')
                                  for object_name in object_names:
                                      if not (object_name.hasAttribute('object_type')):
                                          nombreAlias1 = object_name.firstChild.nodeValue
                                          diccionarioTablas.append((nombreAlias1,nombreTabla))
                                          #print("Nombre Alias: "+nombreAlias1)
                          except Exception as ecc:
                              pass
                              #print(ecc)                              
                ##SELECT##
                #print("## Imprimiendo SELECT ##")
                selects = xmldoc.getElementsByTagName('select_list') 
                for select in selects:
                    lines = select.getElementsByTagName('result_column')
                    for line in lines:
                        
                        expresion = line.getElementsByTagName('expression')[0]
                        tipoExpresion = expresion.attributes['expr_type'].value
                        #if ("simple_object_name_t" in tipoExpresion):
                        objectNames = line.getElementsByTagName('objectName')
                        for objectName in objectNames:
                            nombreOrigen=""
                            nombreAlias=""
                            if (objectName.hasAttribute('object_type') and objectName.attributes['object_type'].value=="column"):
                                try:
                                    #Procesar esto primero porque lo segundo puede fallar
                                    fullName=objectName.getElementsByTagName('part_name')[0]
                                    #print("--")
                                    #if (fullName.firstChild.nodeValue is not "*"):
                                    nombreOrigen=fullName.firstChild.nodeValue
                                    #print("Nombre Origen: "+nombreOrigen)
                                    ########
                                    tableName=objectName.getElementsByTagName('object_name')[0]
                                    tablaOrigen=tableName.firstChild.nodeValue
                                    for dupla in diccionarioTablas:
                                          if (tablaOrigen==dupla[0]):
                                              tablaOrigen=dupla[1]
                                              #print("Nombre TablaX: "+dupla[1])
                                    #print("Tabla Origen: "+tablaOrigen)
                                except Exception as exc:
                                        tablaOrigen=tablaPrincipal
                                        #print("Tabla Origen: "+tablaOrigen)
                                    
                        try:
                              aliasNames=line.getElementsByTagName('alias_clause')
                              for aliasName in aliasNames:
                                  object_names=aliasName.getElementsByTagName('object_name')
                                  for object_name in object_names:
                                      if not (object_name.hasAttribute('object_type')):
                                          nombreAlias = object_name.firstChild.nodeValue
                                          #print("Nombre Alias: "+nombreAlias)
                        except Exception as ecc:
                              print(ecc)

                        # re-open and append
                        if (nombreOrigen):
                            try:
                                wb = load_workbook(os.path.join(root, root+".xlsx"))
                            except:
                                wb = open_xls_as_xlsx(os.path.join(root, root+".xls"))
                            ws = wb.active
                            if (nombreAlias):
                                ws.append(["Extraccion", "Query", "Query Alias","Query",nombreOrigen,nombreAlias])
                            else:
                                ws.append(["Extraccion", "Query", "Query Alias","Query",nombreOrigen,nombreOrigen])
                            wb.save(os.path.join(root, root+".xlsx"))
                                            
                        if (tablaOrigen and len(tablaOrigen)>3):
                            try:
                                wb = load_workbook(os.path.join(root, root+".xlsx"))
                            except:
                                wb = open_xls_as_xlsx(os.path.join(root, root+".xls"))
                            ws = wb.active
                            ws.append(["Extraccion", "Query", "Query Source Table","Query",tablaOrigen,nombreOrigen])
                            wb.save(os.path.join(root, root+".xlsx"))
        gc.collect()
        ##ojo nuevo
        if not contieneXML:
            wb = open_xls_as_xlsx(os.path.join(root, root+".xls"))
            wb.save(os.path.join(root, root+".xlsx"))
            
        


                        
                                  
                              
##  Asociar registros como append al excel actual                    
##  Insertar tabla fuente como default                        
##  Formatear Archivo final
##  Pasar todos los query por la pagina web

                              
##table_reference                          
##namedReferences = fromx.getElementsByTagName('named_table_reference')
##                    
                
                                
                                
                                                
                        
                    
                    
    
        
